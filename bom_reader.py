# BOM reader for MISys

import os
import pyautogui
import time
import openpyxl

pyautogui.PAUSE = 0.5

print('Press Ctrl-C to quit.')

# new line (925, 330)
# IDLE (1890, 1006)

# ask for fine name and complete the path automatically

job_number = input('Whats the job number?: ').upper()
os.chdir('C:\\Users\\jlee.NTPV\\Documents\\BOM')
for i in os.listdir():
    print(i)
    i_list = i.split(' ')
    if job_number == i_list[0]:
        path = 'C:\\Users\\jlee.NTPV\\Documents\\BOM\\' + i

# see how many rows
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row

placeholder = []
num_to_skip = []
job_range = []
stop_loop = False
skip_me = str(input("Do you have numbers to skip? (Y/N): "))
print('Time Start.')
start_time = time.time()
if 'Y' in skip_me or 'y' in skip_me:
    # stop_loop is a secondary measure to prevent infinite loops, not required, but precautionary
    while not stop_loop:
        user_input = input("Please enter the number you would like to skip (enter + to quit): ")
        try:
            if '+' in str(user_input):
                stop_loop = True
                break
        except ValueError:
            continue
            
        try:
            placeholder.append(int(user_input))
        except ValueError:
            print("Please enter a valid number or + to quit")
            continue
    # We need to remove possible duplicates
    for num in placeholder:
        if num not in num_to_skip:
            num_to_skip.append(num)
    
    # Now we want to build a disjointed list to make the future for loop 1000 times easier

    temp_range = range(3,max_row+1,1)
    disjointer_a = [number for number in num_to_skip if number not in temp_range]
    disjointer_b = [number for number in temp_range if number not in num_to_skip]
    
    # Combining the two lists to make the completed iteration
    job_range = disjointer_a + disjointer_b

# Just checking to see if it's empty, that way we won't error out in future
if not job_range:
    job_range = range(3,max_row+1,1)

# create a loop using the number of rows starting at 3
try:
    for i in job_range:

        part_no = sheet_obj.cell(row = i, column = 5)
        qty = sheet_obj.cell(row = i, column = 2)
        multi = sheet_obj.cell(row = i, column = 3)

        part = str(part_no.value)
        
        if " " in part:
            part = part.replace(" ","")
        if "\n" in part:
            part = part.replace("\n","")
        if "N/A" in part:
            continue
        if "NA" in part and len(part) == 2:
            continue
        if len(part) == 0:
            continue
        if 'BY CUSTOMER' in part or 'BYCUSTOMER' in part:
            continue

        # multiply the MULTI and QTY
        total = qty.value*multi.value
        required = str(total)

        # new line
        pyautogui.click(925, 330)
        # pyautogui.click(919,313)
        time.sleep(0.5)

        # type out part_no
        pyautogui.typewrite(part)

        # tab
        pyautogui.typewrite(['tab'])
        time.sleep(0.1)

        # type required
        pyautogui.typewrite(str(required))
        time.sleep(0.1)

        print(required, part)

except KeyboardInterrupt:
    print('\nDone')

elapsed_time = round(time.time()-start_time, 3)
print('\nElapsed time: ' + str(elapsed_time))
print('\nComplete.')

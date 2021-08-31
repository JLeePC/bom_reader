import os
import time
import openpyxl
import pyperclip
import pandas as pd
import datetime as dt
import pygetwindow as gw
import pyautogui as pygui
from decimal import Decimal
from modules.coordinates import *
from modules.get_bom_data import get_bom_data
from PIL import Image, ImageGrab, ImageFilter

class WrongItem(Exception):
    pass

# just click the terminal to be able to type in it if its not already focused
def focus_terminal():
    # i can tell if its already focused if the mouse is already there
    mouse_postion = pygui.position()
    if not mouse_postion.x == terminal[0] and not mouse_postion.y == terminal[1]:
        pygui.click(terminal[0], terminal[1])

# for copying what text is highlighted to save as a variable 
def copy_clipboard():
    pygui.hotkey('ctrl', 'c')
    time.sleep(.01)
    return pyperclip.paste()

# when incorrect part is entered a window pops up. detect that to tell the user to fix the part and get rid of the window
def item_doesnt_exist(x,y):
    print(f'Item {x} does not exist in {y}. Please fix then try again.')
    pygui.click(new_no)
    while len(gw.getWindowsWithTitle('Bills of Material - North Texas Pressure Vessels Inc. Error')) == 0:
        time.sleep(1)
    pygui.click(new_ok)
    while len(gw.getWindowsWithTitle('Bills of Material - North Texas Pressure Vessels Inc. Error')) == 1:
        time.sleep(1)
    pygui.click(refresh)
    while len(gw.getWindowsWithTitle('Bills of Material - North Texas Pressure Vessels Inc.')) == 1:
        time.sleep(1)
    pygui.click(refresh_yes)

def main():
    # see if the 'Bill of Materials.CSV' file exists and is recent. if its not export the data again
    today = dt.datetime.now().date()
    try:
        filetime = dt.datetime.fromtimestamp(os.path.getctime('D:\MIsys Data\Bill of Materials\Bill of Material Details.CSV'))

    except FileNotFoundError:
        print("'Bill of Material Details.CSV' File not found")
        get_bom_data()
        filetime = dt.datetime.fromtimestamp(os.path.getctime('D:\MIsys Data\Bill of Materials\Bill of Material Details.CSV'))

    if filetime.date() != today:
        os.remove('D:\MIsys Data\Bill of Materials\Bill of Material Details.CSV')
        get_bom_data()
        main()

    else:
        # loop through boms in bom folder
        for filename in os.listdir(r"C:\Users\jlee.NTPV\Documents\BOM"):
            if filename.endswith(".xlsx"):
                # load xlsx and read the lines to store the bom
                wb_obj = openpyxl.load_workbook("C:\\Users\\jlee.NTPV\\Documents\\BOM\\"+filename)
                drawing_no = filename.replace(".xlsx","")
                drawing_no_list = drawing_no.split(" ")
                excel_rev = drawing_no_list[2]
                excel_rev = int(excel_rev.replace("R",""))
                excel_qty = drawing_no_list[3]
                excel_qty = int(excel_qty.replace("X",""))
                sheet_obj = wb_obj.active
                if 'BILL OF MATERIAL' in sheet_obj.cell(row=1, column=1).value:
                    start_range = 3
                else:
                    start_range = 2
                max_row = sheet_obj.max_row
                active_excel = []
                for i in range(start_range,max_row+1,1):
                    builder = {}
                    part_no = sheet_obj.cell(row = i, column = 5)
                    qty = sheet_obj.cell(row = i, column = 2)
                    multi = sheet_obj.cell(row = i, column = 3)
                    part_no = str(part_no.value)
                    if " " in part_no:
                        part_no = part_no.replace(" ","")
                    if "\n" in part_no:
                        part_no = part_no.replace("\n","")
                    if "N/A" in part_no:
                        continue
                    if "NA" in part_no and len(part_no) == 2:
                        continue
                    if len(part_no) == 0:
                        continue
                    if 'BY CUSTOMER' in part_no or 'BYCUSTOMER' in part_no:
                        continue

                    # multiply the MULTI and QTY
                    total = Decimal(qty.value*multi.value*excel_qty)
                    total = float(round(total,3))

                    # add info to a list
                    builder['part'] = part_no
                    builder['qty'] = total
                    active_excel.append(builder)

                # input the active drawing number into misys and wait for it to load
                pygui.doubleClick(item_no)
                pygui.typewrite(drawing_no_list[0])
                pygui.click(header)
                time.sleep(1)
                if len(gw.getWindowsWithTitle('Bills of Material - North Texas Pressure Vessels Inc.')) == 2:
                    print(f'{drawing_no_list[0]} does not exist')
                    pygui.click(item_no_ok)
                    time.sleep(1)
                    pygui.click(item_no_ok)
                    time.sleep(1)
                    continue
                os.chdir('png')
                while pygui.locateOnScreen('allocated.png', region=(42,190,120,20)) is None:
                    time.sleep(1)
                os.chdir('..')
                pygui.click(material)
                os.chdir('png')
                while pygui.locateOnScreen('alternates.png', region=(124,1000,75,25)) is None:
                    time.sleep(1)
                os.chdir('..')

                # make sure the bom rev equals the rev from the drawing bom
                pygui.doubleClick(rev_no)
                rev = int(copy_clipboard())
                data = pd.read_csv('D:\MIsys Data\Bill of Materials\Bill of Material Details.CSV')
                if rev == excel_rev:
                    active_bom = []
                    data_drop = []
                    # loop through the data and pick out the applicable lines
                    for i in range(len(data)):
                        builder = {}
                        data_bom = data['bomItem'][i]
                        try:
                            data_rev = int(data['bomRev'][i])
                        except ValueError:
                            continue
                        if drawing_no_list[0] == data_bom and excel_rev == data_rev:
                            data_drop.append(i)
                            builder['line'] = data['lineNbr'][i]
                            builder['part'] = data['partId'][i]
                            builder['qty'] = data['qty'][i]
                            active_bom.append(builder)

                    # if list is empty that means the rev doesnt exist
                    if len(active_bom) == 0:
                        get_bom_data()
                        main()

                    # reorder the data so it will match the excel bom
                    ordered_bom = []
                    oredered_bom_items = []
                    for line in range(0,len(active_bom)+2):
                        for i in active_bom:
                            if int(i['line']) == line - 1:
                                ordered_bom.append(i)
                                if 'LABOR-' not in i['part']:
                                    builder = {}
                                    builder['part'] = i['part']
                                    builder['qty'] = i['qty']
                                    oredered_bom_items.append(builder)

                    # if they dont match delete all rows in misys and input the correct data
                    if oredered_bom_items == active_excel:
                        print(f"{drawing_no_list[0]} matched MiSys.")
                    else:
                        difference = [item for item in active_excel if item not in oredered_bom_items]
                        print(difference)
                        print(f'The first different item was {difference[0]}')
                        try:
                            pygui.doubleClick(item_no)
                            pygui.click(delete_line,clicks=len(ordered_bom))
                            new_bom = []
                            line_no = 1
                            for row in active_excel:
                                builder = {}
                                pygui.click(new_line)
                                pygui.typewrite(row['part'])
                                pygui.typewrite(['tab'])
                                time.sleep(1)
                                if len(gw.getWindowsWithTitle(f'New BOM Component for BOM No. {drawing_no_list[0]} Rev {excel_rev}')) == 1:
                                    raise WrongItem
                                pygui.typewrite(str(row['qty']))
                                builder['RECTYPE'] = "<TBLNAME>"
                                builder['MIBOMD'] = "MIBOMD"
                                builder['bomItem'] = drawing_no_list[0]
                                builder['bomRev'] = str(excel_rev)
                                builder['bomEntry'] = line_no
                                builder['lineNbr'] = line_no
                                builder['dType'] = 0
                                builder['partId'] = row['part']
                                builder['qty'] = row['qty']
                                builder['lead'] = 0
                                builder['cmnt'] = ""
                                builder['opCode'] = ""
                                builder['srcLoc'] = "NTPV"
                                builder['altItems'] = 0

                                new_bom.append(builder)

                                line_no = line_no + 1

                            for row in ordered_bom:
                                builder = {}
                                if 'LABOR-' in row['part']:
                                    pygui.click(new_line)
                                    pygui.typewrite(row['part'])
                                    pygui.typewrite(['tab'])
                                    time.sleep(1)
                                    pygui.typewrite(str(row['qty']))

                                    # builder['RECTYPE'] = "<TBLNAME>"
                                    # builder['MIBOMD'] = "MIBOMD"
                                    builder['bomItem'] = drawing_no_list[0]
                                    builder['bomRev'] = str(excel_rev)
                                    # builder['bomEntry'] = line_no
                                    builder['lineNbr'] = line_no
                                    # builder['dType'] = 0
                                    builder['partId'] = row['part']
                                    builder['qty'] = row['qty']
                                    # builder['lead'] = 0
                                    # builder['cmnt'] = ""
                                    # builder['opCode'] = ""
                                    # builder['srcLoc'] = "NTPV"
                                    # builder['altItems'] = 0

                                    new_bom.append(builder)
                                    
                                    line_no = line_no + 1

                        except WrongItem:
                            item_doesnt_exist(row['part'],drawing_no_list[0])

                        # edit data to match what is now in misys
                        rows = data.index[data_drop]
                        data.drop(rows, inplace=True)
                        for row in new_bom:
                            data = data.append(row,True)
                        os.remove('D:\MIsys Data\Bill of Materials\Bill of Material Details.CSV')
                        data.to_csv('D:\MIsys Data\Bill of Materials\Bill of Material Details.CSV', index=False)

                    # save bom
                    pygui.click(save)

                    # delete excel bom from folder
                    os.remove("C:\\Users\\jlee.NTPV\\Documents\\BOM\\"+filename)

                elif rev < excel_rev:
                    print("{} rev does not match.".format(drawing_no_list[0]))
                    pygui.click(save)
                    pygui.click(new_rev)
                    while len(gw.getWindowsWithTitle('Bills of Material - North Texas Pressure Vessels Inc.')) == 1:
                        time.sleep(1)
                    pygui.click(new_rev_yes)
                    while len(gw.getWindowsWithTitle('Bills of Material - North Texas Pressure Vessels Inc.')) > 1:
                        time.sleep(1)
                    pygui.doubleClick(rev_no)
                    pygui.typewrite(str(excel_rev))
                    pygui.click(revision)
                    os.chdir('png')
                    while pygui.locateOnScreen('revision_date.png', region=(42,190,96,20)) is None:
                        time.sleep(1)
                    os.chdir('..')
                    pygui.click(current_revision)
                    pygui.click(material)
                    os.chdir('png')
                    while pygui.locateOnScreen('alternates.png', region=(124,1000,75,25)) is None:
                        time.sleep(1)
                    os.chdir('..')
                    pygui.click(save)

                    # add new rev to csv
                    active_bom = []
                    # loop through the data and pick out the applicable lines
                    for i in range(len(data)):
                        builder = {}
                        data_bom = data['bomItem'][i]
                        try:
                            data_rev = int(data['bomRev'][i])
                        except ValueError:
                            continue
                        if drawing_no_list[0] == data_bom and excel_rev == data_rev:
                            builder['line'] = data['lineNbr'][i]
                            builder['part'] = data['partId'][i]
                            builder['qty'] = data['qty'][i]
                            active_bom.append(builder)

                    # reorder the data so it will match the excel bom
                    ordered_bom = []
                    for line in range(0,len(active_bom)+2):
                        for i in active_bom:
                            if int(i['line']) == line - 1:
                                ordered_bom.append(i)
                                if 'LABOR-' not in i['part']:
                                    builder = {}
                                    builder['part'] = i['part']
                                    builder['qty'] = i['qty']

                    data = data.append(row,True)
                    os.remove('D:\MIsys Data\Bill of Materials\Bill of Material Details.CSV')
                    data.to_csv('D:\MIsys Data\Bill of Materials\Bill of Material Details.CSV', index=False)

                    main()
                
                else:
                    print('Bom rev is lower than MiSys rev.')
                    continue

if __name__ == '__main__':
    main()

focus_terminal()

print("Done.")
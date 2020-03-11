import os
import time
import openpyxl
import pyautogui
import pyperclip
import pytesseract
import pygetwindow as gw
from PIL import Image, ImageGrab, ImageFilter

pytesseract.pytesseract.tesseract_cmd = r'C:\Users\jlee.NTPV\AppData\Local\Tesseract-OCR\tesseract.exe'
custom_oem_psm_config = r'--oem 3 --psm 6'

item_no = 206, 92
rev = 454, 94
save = 77, 59
header = 42, 161
material = 205, 161
item = 304, 249
new_line = 1885, 320
trash = 1884, 342
qty = 591, 250
terminal = 2385, 15

def remove_values_from_list(the_list, val):
   return [value for value in the_list if value != val]

def copy_clipboard():
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(.01)
    return pyperclip.paste()

def go_up():
    if scroll_bar_switch:
        os.chdir("png")
        up_arrow_png = pyautogui.locateOnScreen('up_arrow.png', region=(up_arrow_x1,up_arrow_y1,20,20))
        top_arrow = pyautogui.locateOnScreen('top_arrow.png', region=(up_arrow_x1,up_arrow_y1,20,40))
        if up_arrow_png is not None:
            if top_arrow is None:
                pyautogui.click(up_arrow_x_center, up_arrow_y_center + 15, clicks=3, interval=0.1)
                pyautogui.click(up_arrow_x_center, up_arrow_y_center)
                pyautogui.moveTo(qty[0], qty[1])
                while top_arrow is None:
                    top_arrow = pyautogui.locateOnScreen('top_arrow.png', region=(up_arrow_x1,up_arrow_y1,20,40))
                    time.sleep(0.1)
                time.sleep(0.2)
        os.chdir("..")

def screen_reader(x1,y1,x2,y2):
    image = ImageGrab.grab(bbox =(x1,y1,x2,y2))
    new_size = tuple(4*x for x in image.size)
    image = image.resize(new_size, Image.ANTIALIAS)
    image_bl = image.filter(ImageFilter.GaussianBlur(radius = 1))
    image_gs = image_bl.convert('LA')
    image_string = pytesseract.image_to_string(image_gs, config=custom_oem_psm_config)
    image_string_list = image_string.split("\n")
    result = remove_values_from_list(image_string_list, "")
    return(result)

def arrow_up(x):
    for _ in range(x):
        pyautogui.typewrite(['up'])
    wait = x / 25
    if wait < .25:
        wait = .25
    time.sleep(wait)

def arrow_down(x):
    for _ in range(x):
        pyautogui.typewrite(['down'])
    wait = x / 25
    if wait < .25:
        wait = .25
    time.sleep(wait)

def focuse_terminal():
    mouse_postion = pyautogui.position()
    if not mouse_postion.x == terminal[0] and not mouse_postion.y == terminal[1]:
        pyautogui.click(terminal[0], terminal[1])
job_loop = True

# ------------------------------------------------------------------

os.chdir('C:\\Users\\jlee.NTPV\\Documents\\BOM')

for file in os.listdir():
    file_list = file.split(' ')
    bom = file_list[0]
    rev = file_list[1]
    rev = rev.replace('R', '')
    rev = rev.replace('.xlsx', '')

    pyautogui.doubleClick(item_no[0], item_no[1])
    pyautogui.typewrite(bom)
    pyautogui.typewrite(['tab'])
    pyautogui.click(header[0], header[1])
    os.chdir(r"C:\Users\jlee.NTPV\Documents\GitHub\bom_reader\png")
    try:
        up_arrow_x_center, up_arrow_y_center = pyautogui.locateCenterOnScreen('up_arrow.png')
        up_arrow_x1 = up_arrow_x_center - 10
        up_arrow_y1 = up_arrow_y_center - 10
        down_arrow_x_center, down_arrow_y_center = pyautogui.locateCenterOnScreen('down_arrow.png')
        scroll_bar_switch = True
    except TypeError:
        scroll_bar_switch = False

    os.chdir("..")
    pyautogui.click(qty[0], qty[1])# click HT# box
    pyautogui.typewrite(['up'])
    pyautogui.typewrite(['up'])
    go_up()
    # read all part no. to know if there are labor or pwht lines and where they are located
    read_switch = True
    pwht_in_mo = False
    labor_in_mo = False
    os_in_mo = False
    mo_list = []
    line_number_list = []
    while read_switch:
        # parts = screen_reader(part_no_x1,part_no_y1,part_no_x2,part_no_y2)
        # if I use the second two corrdinates from the part_no it knows when the next line is and reads it better
        lines = screen_reader(line_number_x1,line_number_y1,part_no_x2,part_no_y2)
        if len(lines) == 0:
            lines = screen_reader(line_number_x1,line_number_y1,part_no_x2,line_number_y3)
        line = lines[len(lines)-1]
        line_list = line.split(" ")
        line = line_list[0]
        if "," in line:
            line = line.replace(",", "")
        if int(line) > len(mo_list):
            for i in range(0,len(lines)):
                builder = {}
                line = lines[i]
                line_list = line.split(" ")
                line = line_list[0]
                part = line_list[1]
                try:
                    builder['Line'] = int(line)
                except ValueError:
                    for j in range(len(line)):
                        line = j
                    builder['Line'] = int(line)
                builder['Part No.'] = part
                mo_list.append(builder)
            if len(lines) < 35:
                read_switch = False
            else:
                if scroll_bar_switch:
                    pyautogui.click(down_arrow_x_center, down_arrow_y_center, clicks=35, interval=0.1) # down arrow
                else:
                    read_switch = False
        else:
            read_switch = False
    
    for j in mo_list:
        if 'LABOR' in j['Part No.']:
            j['Labor'] = True
            labor_in_mo = True
        else:
            j['Labor'] = False
        if 'PWHT' in j['Part No.']:
            j['PWHT'] = True
            pwht_in_mo = True
        else:
            j['PWHT'] = False
        if 'OS-' in j['Part No.']:
            j['OS-'] = True
            os_in_mo = True
        else:
            j['OS-'] = False

    path = 'C:\\Users\\jlee.NTPV\\Documents\\BOM\\' + file

    # see how many rows
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

    job_range = range(3,max_row+1,1)

    # create a loop using the number of rows starting at 3
    for i in job_range:

        part_no = sheet_obj.cell(row = i, column = 5)
        qty = sheet_obj.cell(row = i, column = 2)
        multi = sheet_obj.cell(row = i, column = 3)

        part = str(part_no.value)
        
        if " " in part:
            part = part.replace(" ","")
        if "\n" in part:
            part = part.replace("\n","")
        if "N/A" in part:
            continue
        if len(part) == 0:
            continue

        # multiply the MULTI and QTY
        total = qty.value*multi.value
        required = str(total)

        # new line
        pyautogui.click(925, 330)
        #pyautogui.click(919,313)

        # type out part_no
        pyautogui.typewrite(part)
        time.sleep(0.1)

        # tab
        pyautogui.typewrite(['tab'])

        # type required
        pyautogui.typewrite(str(required))
        time.sleep(0.1)